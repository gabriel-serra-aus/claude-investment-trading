"""
load_stock_prices.py
Loads multi-tab stock price Excel into a flat Stock Prices table.

Usage:
    python load_stock_prices.py \
        --source  <path/to/USD Stock Prices History.xlsx> \
        --target  <path/to/Portfolio.xlsx> \
        --sheet   "Stock Prices" \
        --table   "Table4"
"""
import argparse
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import TableStyleInfo

FMT_DATE     = "mm-dd-yy"
FMT_CURRENCY = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

def find_col_indices(ws):
    """Return (date_col, price_col) 1-based indices from header row."""
    date_col = price_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip().lower()
        if h == "date":
            date_col = c
        elif h in ("close", "price", "usd price", "adj close"):
            price_col = c
    return date_col, price_col

def load(source_path, target_path, sheet_name="Stock Prices", table_name=None):
    # ── Read source ──────────────────────────────────────────────────────
    src = load_workbook(source_path, data_only=True, read_only=True)
    all_rows = []

    for ticker in src.sheetnames:
        ws = src[ticker]
        date_col, price_col = find_col_indices(ws)
        if date_col is None or price_col is None:
            print(f"  WARNING: {ticker} — could not identify Date/Price columns, skipping")
            continue
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            date_val = row[date_col - 1]
            price    = row[price_col - 1]
            if date_val is None or price is None:
                continue
            if isinstance(date_val, datetime):
                date_val = date_val.date()
            all_rows.append((date_val, ticker, price))
            count += 1
        print(f"  {ticker}: {count} rows")
    src.close()

    total = len(all_rows)
    print(f"Total rows: {total}")
    if total == 0:
        print("Nothing to write.")
        return

    all_rows.sort(key=lambda x: (x[1], x[0]))

    # ── Write to target ──────────────────────────────────────────────────
    wb = load_workbook(target_path)
    ws_t = wb[sheet_name]

    # Find table
    if table_name is None:
        table_name = list(ws_t.tables.keys())[0] if ws_t.tables else None
    if table_name is None:
        raise ValueError(f"No table found on sheet '{sheet_name}'")

    # Clear existing data rows
    for r in range(2, ws_t.max_row + 1):
        for c in range(1, 5):
            ws_t.cell(r, c).value = None

    # Check if Rates sheet exists for AUD formula
    has_rates = "Rates" in wb.sheetnames

    for i, (date_val, ticker, usd_price) in enumerate(all_rows, start=2):
        ws_t.cell(i, 1, date_val).number_format      = FMT_DATE
        ws_t.cell(i, 2, ticker)
        ws_t.cell(i, 3, usd_price).number_format     = FMT_CURRENCY
        if has_rates:
            c = ws_t.cell(i, 4)
            c.value          = f'=IFERROR(C{i}/INDEX(Rates!$B:$B,MATCH(A{i},Rates!$A:$A,0)),"")'
            c.number_format  = FMT_CURRENCY

    # Update table ref
    last_row = total + 1
    tbl = ws_t.tables[table_name]
    tbl.ref = f"A1:D{last_row}"
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,   showColumnStripes=False,
    )
    print(f"Table '{table_name}' ref → A1:D{last_row}")

    wb.save(target_path)
    print(f"Saved: {target_path}")

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--source",  required=True)
    p.add_argument("--target",  required=True)
    p.add_argument("--sheet",   default="Stock Prices")
    p.add_argument("--table",   default=None)
    args = p.parse_args()
    load(args.source, args.target, args.sheet, args.table)
